"""Lenient Excel import and two-sheet Plan export."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from .domain import DomainValidationError, Plan, Task, next_task_id, parse_date


def _header_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


COLUMN_ALIASES = {
    "name": {"задача", "task", "name", "название", "наименование"},
    "description": {"описание", "description", "details", "детали"},
    "assignee": {"исполнитель", "assignee", "owner", "ответственный"},
    "duration": {"длительность", "duration", "days", "дни"},
    "predecessors": {"предшественники", "predecessors", "predecessor", "dependencies"},
    "pinned_start": {"привязка", "pinned start", "pinned_start", "start pin"},
    "due_date": {"срок", "due date", "due_date", "deadline"},
    "id": {"id", "task id", "идентификатор"},
    "plan_start": {"начало плана", "plan start", "plan_start"},
    "plan_name": {"название плана", "plan name", "plan_name"},
}
ALIAS_TO_FIELD = {alias: field for field, aliases in COLUMN_ALIASES.items() for alias in aliases}


@dataclass
class ImportReport:
    total_rows: int = 0
    loaded_count: int = 0
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    found_headers: List[str] = field(default_factory=list)

    @property
    def rejected_count(self) -> int:
        return max(0, self.total_rows - self.loaded_count)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "total_rows": self.total_rows,
            "loaded_count": self.loaded_count,
            "rejected_count": self.rejected_count,
            "warnings": self.warnings,
            "errors": self.errors,
            "found_headers": self.found_headers,
            "summary": f"Загружено {self.loaded_count} задач из {self.total_rows}",
        }


def _parse_duration(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        numeric = float(str(value).strip())
        if not numeric.is_integer():
            return None
        result = int(numeric)
        return result if result >= 1 else None
    except (TypeError, ValueError):
        return None


def _split_predecessors(value: Any) -> List[str]:
    if value in (None, ""):
        return []
    return list(dict.fromkeys(part.strip() for part in re.split(r"[,;\n]+", str(value)) if part.strip()))


def _resolve_predecessors(tasks: Dict[str, Task], raw: Dict[str, List[str]], report: ImportReport) -> None:
    by_name: Dict[str, List[str]] = {}
    for task in tasks.values():
        by_name.setdefault(task.name.strip().casefold(), []).append(task.id)
    for task_id, references in raw.items():
        resolved: List[str] = []
        for reference in references:
            if reference in tasks:
                resolved.append(reference)
                continue
            candidates = by_name.get(reference.casefold(), [])
            if len(candidates) == 1:
                resolved.append(candidates[0])
            elif len(candidates) > 1:
                report.warnings.append(f"Связь для «{tasks[task_id].name}» по имени «{reference}» не добавлена: имя неоднозначно")
            else:
                report.warnings.append(f"Предшественник «{reference}» для «{tasks[task_id].name}» не существует и пропущен")
        tasks[task_id].predecessors = list(dict.fromkeys(resolved))


def _break_cycles(tasks: Dict[str, Task], report: ImportReport) -> None:
    """Remove the DFS back-edge, one at a time, until the imported graph is valid."""
    while True:
        state: Dict[str, int] = {task_id: 0 for task_id in tasks}
        stack: List[str] = []
        removed: Optional[Tuple[str, str]] = None

        def visit(task_id: str) -> None:
            nonlocal removed
            if removed:
                return
            state[task_id] = 1
            stack.append(task_id)
            for predecessor_id in list(tasks[task_id].predecessors):
                if state[predecessor_id] == 1:
                    tasks[task_id].predecessors.remove(predecessor_id)
                    removed = (task_id, predecessor_id)
                    return
                if state[predecessor_id] == 0:
                    visit(predecessor_id)
                    if removed:
                        return
            stack.pop()
            state[task_id] = 2

        for task_id in tasks:
            if state[task_id] == 0:
                visit(task_id)
            if removed:
                break
        if not removed:
            return
        task_id, predecessor_id = removed
        report.warnings.append(f"Цикл разорван: удалена замыкающая связь «{tasks[task_id].name}» → «{tasks[predecessor_id].name}»")


def import_plan(data: bytes, default_plan_start: str = "2026-09-01") -> Tuple[Plan, ImportReport]:
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    report = ImportReport()
    if not rows:
        report.errors.append("Файл не содержит строки заголовков")
        raise ValueError(report.errors[0])

    raw_headers = list(rows[0])
    report.found_headers = [str(value or "") for value in raw_headers]
    fields: Dict[str, int] = {}
    for index, raw_header in enumerate(raw_headers):
        field_name = ALIAS_TO_FIELD.get(_header_key(raw_header))
        if field_name and field_name not in fields:
            fields[field_name] = index
    if "name" not in fields or "duration" not in fields:
        expected = "задача, описание, исполнитель, длительность, предшественники"
        raise ValueError(f"Не удалось распознать обязательные заголовки. Найдены: {', '.join(report.found_headers)}. Ожидались: {expected}")

    data_rows = [row for row in rows[1:] if any(value not in (None, "") for value in row)]
    report.total_rows = len(data_rows)
    tasks: Dict[str, Task] = {}
    raw_predecessors: Dict[str, List[str]] = {}
    plan_start_index = fields.get("plan_start")
    raw_plan_start: Any = (
        data_rows[0][plan_start_index] if data_rows and plan_start_index is not None and plan_start_index < len(data_rows[0]) else None
    )

    for row in data_rows:

        def value(field_name: str, fallback: Any = "") -> Any:
            index = fields.get(field_name)
            return row[index] if index is not None and index < len(row) else fallback

        name = str(value("name") or "").strip()
        duration = _parse_duration(value("duration"))
        if not name:
            report.errors.append("Строка без названия задачи пропущена")
            continue
        if duration is None:
            report.errors.append(f"Задача «{name}»: длительность пустая или нечисловая")
            continue
        task_id = str(value("id") or next_task_id(tasks)).strip()
        if task_id in tasks:
            task_id = next_task_id(tasks)
            report.warnings.append(f"Дублирующийся ID для «{name}» заменён на «{task_id}»")
        pinned_start = value("pinned_start")
        due_date = value("due_date")
        try:
            parsed_pinned = parse_date(pinned_start)
        except DomainValidationError:
            parsed_pinned = None
            report.errors.append(f"Задача «{name}»: некорректная дата привязки")
        try:
            parsed_due = parse_date(due_date)
        except DomainValidationError:
            parsed_due = None
            report.errors.append(f"Задача «{name}»: некорректный срок")
        tasks[task_id] = Task(
            id=task_id,
            name=name,
            description=value("description") or "",
            assignee=value("assignee") or "",
            duration=duration,
            pinned_start=parsed_pinned,
            due_date=parsed_due,
        )
        raw_predecessors[task_id] = _split_predecessors(value("predecessors"))
        report.loaded_count += 1

    if not tasks:
        raise ValueError("Файл не содержит ни одной задачи с корректной длительностью")
    _resolve_predecessors(tasks, raw_predecessors, report)
    _break_cycles(tasks, report)
    plan_name = "Импортированный план"
    if data_rows and fields.get("plan_name") is not None:
        first_value = data_rows[0][fields["plan_name"]]
        if first_value not in (None, ""):
            plan_name = str(first_value).strip()
    parsed_plan_start = parse_date(default_plan_start)
    if raw_plan_start not in (None, ""):
        try:
            parsed_plan_start = parse_date(raw_plan_start) or parsed_plan_start
        except DomainValidationError:
            report.errors.append("Некорректная дата начала плана; использовано значение по умолчанию")
    plan = Plan("imported", plan_name, parsed_plan_start, tasks)
    plan.schedule()
    return plan, report


def _predecessor_text(task: Task, plan: Plan) -> str:
    counts: Dict[str, int] = {}
    for item in plan.tasks.values():
        counts[item.name] = counts.get(item.name, 0) + 1
    return ", ".join(
        plan.tasks[predecessor_id].id if counts[plan.tasks[predecessor_id].name] > 1 else plan.tasks[predecessor_id].name
        for predecessor_id in task.predecessors
    )


def export_plan(plan: Plan) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План"
    sheet.append(
        [
            "Задача",
            "Описание",
            "Исполнитель",
            "Длительность",
            "Предшественники",
            "Pinned Start",
            "Due Date",
            "ID",
            "Plan Start",
            "Plan Name",
        ]
    )
    for task in plan.tasks.values():
        sheet.append(
            [
                task.name,
                task.description,
                task.assignee,
                task.duration,
                _predecessor_text(task, plan),
                task.pinned_start,
                task.due_date,
                task.id,
                plan.plan_start,
                plan.name,
            ]
        )

    schedule_sheet = workbook.create_sheet("Расписание")
    schedule_sheet.append(["ID", "Задача", "Начало", "Последний день", "Due Date", "Статус"])
    schedule = plan.schedule()
    for task in plan.tasks.values():
        item = schedule[task.id]
        overdue = task.due_date and item.last_day > task.due_date
        schedule_sheet.append([task.id, task.name, item.start, item.last_day, task.due_date, "Просрочено" if overdue else "В срок"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
